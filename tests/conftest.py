"""Synthetic XTB workbooks.

Built in memory rather than committed as a binary. A real export carries an
account number, every trade and every cash movement, so it has no business in
a repository — and a generated fixture can be bent into the shapes that matter
(a grown preamble, a multi-lot holding, a missing sheet) which a captured file
cannot.

The layout mirrors a real export exactly: a preamble of a different height per
sheet, a summary block above the open-positions table, and a trailing totals
row on both data sheets.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

# 2026-06-02T12:00:00 and 2026-07-02T12:00:00 as Excel serials — thirty days
# apart, so the average holding period is an exact number.
OPENED = 46175.5
CLOSED = 46205.5

OPEN_HEADER = [
    "Product", "Instrument/Position", "Ticker", "Category", "Type", "Volume",
    "Value", "Current price", "Open price", "Open time (UTC)", "Stop Loss",
    "Take Profit", "Net Profit %", "Net Profit", "Gross Profit", "Margin",
    "Open Commission", "Swap", "Rollover",
]

CLOSED_HEADER = [
    "Instrument", "Ticker", "Category", "Type", "Volume", "Open Price",
    "Open Time (UTC)", "Close Price", "Close Time (UTC)", "Product",
    "Profit/Loss", "Gross Profit", "Purchase Value", "Sale Value", "Stop Loss",
    "Take Profit", "Commission", "Margin", "Swap", "Rollover",
    "Open Conversion Rate", "Close Conversion Rate", "Close Origin",
    "Position ID", "Comment",
]

CASH_HEADER = [
    "Type", "Instrument", "Ticker", "Category", "Time", "Amount", "ID",
    "Comment", "Product", "Position ID",
]


def _aggregate(instrument, ticker, category, volume, value, open_price, net_profit):
    """An instrument's summary row: Category filled, Type and Current price blank."""
    row = [None] * len(OPEN_HEADER)
    row[0] = "My Trades"
    row[1] = instrument
    row[2] = ticker
    row[3] = category
    row[4] = ""  # Type — the discriminator
    row[5] = volume
    row[6] = value
    row[7] = ""  # Current price — only the lots carry it
    row[8] = open_price
    row[13] = net_profit
    return row


def _lot(position_id, ticker, side, volume, value, current_price, open_price):
    """One fill under an instrument: Type filled, Category blank."""
    row = [None] * len(OPEN_HEADER)
    row[0] = "My Trades"
    row[1] = position_id
    row[2] = ticker
    row[3] = ""  # Category
    row[4] = side
    row[5] = volume
    row[6] = value
    row[7] = current_price
    row[8] = open_price
    row[9] = OPENED
    return row


def _closed(instrument, ticker, volume, open_price, close_price, profit, position_id):
    row = [None] * len(CLOSED_HEADER)
    row[0] = instrument
    row[1] = ticker
    row[2] = "STOCK"
    row[3] = "BUY"
    row[4] = volume
    row[5] = open_price
    row[6] = OPENED
    row[7] = close_price
    row[8] = CLOSED
    row[10] = profit
    row[23] = float(position_id)
    return row


def _cash(op_type, ticker, amount, at=CLOSED, position_id=0):
    row = [None] * len(CASH_HEADER)
    row[0] = op_type
    row[2] = ticker
    row[3] = "STOCK" if ticker else ""
    row[4] = at
    row[5] = amount
    row[9] = float(position_id) if position_id else None
    return row


DEFAULT_OPEN = [
    # AAA: one lot. Cost 2 x 10 = 20, worth 24, up 4.
    _aggregate("Alpha", "AAA.US", "STOCK", 2.0, 24.0, 10.0, 4.0),
    _lot("1001", "AAA.US", "BUY", 2.0, 24.0, 12.0, 10.0),
    # BBB: two lots folded into one holding. Cost 1 x 50 = 50, worth 45, down 5.
    _aggregate("Beta", "BBB.UK", "ETF", 1.0, 45.0, 50.0, -5.0),
    _lot("1002", "BBB.UK", "BUY", 0.6, 27.0, 45.0, 50.0),
    _lot("1003", "BBB.UK", "BUY", 0.4, 18.0, 45.0, 50.0),
]

DEFAULT_CLOSED = [
    _closed("Gamma", "CCC.US", 1.0, 10.0, 13.0, 3.0, 2001),
    _closed("Gamma", "CCC.US", 1.0, 10.0, 9.0, -1.0, 2002),
]

DEFAULT_CASH = [
    _cash("Deposit", "", 100.0),
    _cash("Stock purchase", "AAA.US", -20.0, position_id=1001),
    _cash("Stock purchase", "BBB.UK", -50.0, position_id=1002),
    _cash("Stock sell", "CCC.US", 13.0, position_id=2001),
    _cash("Stock sell", "CCC.US", 9.0, position_id=2002),
    _cash("Dividend", "AAA.US", 0.5),
    _cash("Withdrawal", "", -10.0),
]

# Sum of DEFAULT_CASH amounts — the figure a real export puts in its Total row.
DEFAULT_CASH_TOTAL = 42.5
DEFAULT_REALIZED = 2.0
DEFAULT_OPEN_VALUE = 69.0
DEFAULT_OPEN_PROFIT = -1.0


def make_workbook(
    *,
    open_rows=None,
    closed_rows=None,
    cash_rows=None,
    extra_preamble: int = 0,
    open_value: float | None = None,
    open_profit: float | None = None,
    realized_total: float | None = None,
    cash_total: float | None = None,
    sheets: tuple[str, ...] = ("Open Positions", "Closed Positions", "Cash Operations"),
) -> bytes:
    """Assemble a workbook. Every default can be swapped for one test's shape."""
    open_rows = DEFAULT_OPEN if open_rows is None else open_rows
    closed_rows = DEFAULT_CLOSED if closed_rows is None else closed_rows
    cash_rows = DEFAULT_CASH if cash_rows is None else cash_rows

    book = Workbook()
    book.remove(book.active)

    if "Open Positions" in sheets:
        sheet = book.create_sheet("Open Positions")
        sheet.append(["Account number", "12345"])
        sheet.append(["Open Positions", ""])
        sheet.append(["Data as of report generated", CLOSED])
        # Extra lines here are the point: a real preamble grows when XTB adds a
        # row, and the parser must find the header anyway.
        for index in range(extra_preamble):
            sheet.append([f"Extra {index}", ""])
        sheet.append(["Product", "Metric", "Amount", "Currency"])
        sheet.append(
            ["My Trades", "Value",
             DEFAULT_OPEN_VALUE if open_value is None else open_value, "USD"]
        )
        sheet.append(
            ["My Trades", "Profit",
             DEFAULT_OPEN_PROFIT if open_profit is None else open_profit, "USD"]
        )
        sheet.append([])
        sheet.append(["Note", "Summary values are shown as of the report generation time"])
        sheet.append(OPEN_HEADER)
        for row in open_rows:
            sheet.append(row)

    if "Closed Positions" in sheets:
        sheet = book.create_sheet("Closed Positions")
        sheet.append(["Account number", "12345"])
        sheet.append(["Closed Positions", ""])
        sheet.append(["Date from (UTC)", OPENED])
        sheet.append(["Date to (UTC)", CLOSED])
        sheet.append(CLOSED_HEADER)
        for row in closed_rows:
            sheet.append(row)
        footer = [None] * len(CLOSED_HEADER)
        footer[0] = "Profit/loss"
        footer[10] = DEFAULT_REALIZED if realized_total is None else realized_total
        sheet.append(footer)

    if "Cash Operations" in sheets:
        sheet = book.create_sheet("Cash Operations")
        sheet.append(["Account number", "12345"])
        sheet.append(["Cash Operations", ""])
        sheet.append(["Date from (UTC)", OPENED])
        sheet.append(["Date to (UTC)", CLOSED])
        sheet.append(CASH_HEADER)
        for row in cash_rows:
            sheet.append(row)
        footer = [None] * len(CASH_HEADER)
        footer[0] = "Total"
        footer[5] = DEFAULT_CASH_TOTAL if cash_total is None else cash_total
        sheet.append(footer)

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def workbook():
    return make_workbook


@pytest.fixture
def isolated_data(tmp_path, monkeypatch):
    """Point every storage path at a temp dir.

    Without this a test run would read — and the sync tests would overwrite —
    the developer's real watchlist.
    """
    from app import config, storage

    paths = {
        "WATCHLIST_FILE": tmp_path / "stocks_config.json",
        "SETTINGS_FILE": tmp_path / "settings.json",
        "RESULTS_CACHE_FILE": tmp_path / "last_results.json",
        "XTB_SNAPSHOT_FILE": tmp_path / "xtb_portfolio.json",
        "WATCHLIST_BACKUP_FILE": tmp_path / "stocks_config.backup.json",
    }
    for name, path in paths.items():
        monkeypatch.setattr(config, name, path, raising=False)
        monkeypatch.setattr(storage, name, path, raising=False)
    return tmp_path
