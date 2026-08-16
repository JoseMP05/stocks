"""Turns an XTB .xlsx account statement into typed models.

Three sheets, three different preambles, and two of them end in a totals row
that is not data. The open-positions sheet nests two kinds of row under a
single header. None of that survives a rectangular read, which is why this
walks cells with openpyxl instead of reaching for pandas.read_excel.

Two decisions here are load-bearing and easy to undo by accident:

* Header rows are located by matching column names, never by index. The
  open-positions sheet carries a summary table, a blank row and a note above
  its real header, and that preamble grows whenever XTB adds a line. An
  off-by-one does not fail — it silently reads every field from the wrong
  column.
* The totals rows are parsed, not just skipped. They are the broker's own
  arithmetic over the same data, which makes them the only outside check that
  the nested rows were folded correctly. `metrics.build_metrics` compares
  against them.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

from app.models import (
    CashBucket,
    XtbCashOperation,
    XtbClosedPosition,
    XtbLot,
    XtbOpenPosition,
    XtbReportedTotals,
    XtbSnapshot,
)
from app.xtb.symbols import to_yahoo

SHEET_OPEN = "Open Positions"
SHEET_CLOSED = "Closed Positions"
SHEET_CASH = "Cash Operations"

# Excel counts days from 1899-12-30, not 1900-01-01: the offset absorbs the
# phantom 1900 leap day Lotus 1-2-3 introduced and Excel kept for compatibility.
EXCEL_EPOCH = datetime(1899, 12, 30)

# Column-A markers on the trailing summary row of a data sheet.
TOTAL_ROW_MARKERS = frozenset({"profit/loss", "total"})

# How far down to look for a header before giving up.
HEADER_SEARCH_LIMIT = 40

CASH_BUCKETS: dict[str, CashBucket] = {
    "stock purchase": "purchase",
    "stock sell": "sell",
    "deposit": "deposit",
    "withdrawal": "withdrawal",
    "dividend": "dividend",
    "withholding tax": "tax",
    "subaccount transfer": "transfer",
}


class XtbParseError(ValueError):
    """A workbook we cannot make sense of.

    Carries a message written for the person who uploaded the file, not a
    validator dump. Raised rather than HTTPException for the same reason as
    PositionInputError in main.py: htmx discards the body of a 4xx, so the
    message has to travel back inside a 200 to be seen at all.
    """


# ── cell helpers ─────────────────────────────────────────────────────────────

def _norm(value: Any) -> str:
    """A cell as trimmed text. None and blanks both become ""."""
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    """A cell as a number, or None when it is blank or not numeric."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _norm(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_id(value: Any) -> str:
    """Format a position ID.

    XTB stores these as numbers, so they arrive as floats — str() would give
    "2614289166.0", which matches nothing.
    """
    number = _to_float(value)
    if number is None:
        return _norm(value)
    return str(int(number))


def excel_serial_to_iso(value: Any) -> str | None:
    """Convert a spreadsheet timestamp to an ISO string.

    Handles both shapes the cell can take: openpyxl converts date-formatted
    cells to datetime itself, but leaves plain numeric ones as serials.
    """
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    serial = _to_float(value)
    if serial is None or serial <= 0:
        return None
    return (EXCEL_EPOCH + timedelta(days=serial)).isoformat(timespec="seconds")


# ── header discovery ─────────────────────────────────────────────────────────

def _find_header(
    rows: list[tuple], required: set[str], sheet: str
) -> tuple[int, dict[str, int]]:
    """Locate a header row by column name and map names to column indices.

    Returns the row's index in `rows` plus {lowercased column name -> index}.
    Raises XtbParseError naming the columns it could not find.
    """
    best_missing: set[str] | None = None
    for index, row in enumerate(rows[:HEADER_SEARCH_LIMIT]):
        labels = {_norm(cell).lower(): position for position, cell in enumerate(row)}
        missing = required - labels.keys()
        if not missing:
            return index, labels
        if best_missing is None or len(missing) < len(best_missing):
            best_missing = missing

    absent = ", ".join(sorted(best_missing or required))
    raise XtbParseError(
        f"La hoja «{sheet}» no tiene el formato esperado: "
        f"faltan las columnas {absent}."
    )


def _cell(row: tuple, columns: dict[str, int], name: str) -> Any:
    index = columns.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _is_total_row(row: tuple, columns: dict[str, int], identity: str) -> bool:
    """True for the summary row a data sheet ends with.

    Two conditions, not one. Column A alone is not enough: a `Deposit` line
    also leaves the instrument column blank, and a future export could name a
    real instrument "Total".
    """
    marker = _norm(row[0]).lower() if row else ""
    return marker in TOTAL_ROW_MARKERS and not _norm(_cell(row, columns, identity))


# ── open positions ───────────────────────────────────────────────────────────

_OPEN_REQUIRED = {"ticker", "category", "type", "volume", "value"}


def _parse_open_summary(rows: list[tuple], header_index: int) -> tuple[XtbReportedTotals, str]:
    """Read the Product/Metric/Amount/Currency block above the position table.

    Amounts are summed across products rather than read off the "My Trades"
    line: an account that also holds Investment Plans reports them separately,
    and both show up in the position table below.
    """
    totals = XtbReportedTotals()
    currency = ""
    value_total: float | None = None
    profit_total: float | None = None

    for row in rows[:header_index]:
        if len(row) < 3:
            continue
        metric = _norm(row[1]).lower()
        amount = _to_float(row[2])
        if amount is None:
            continue
        if metric == "value":
            value_total = (value_total or 0.0) + amount
        elif metric == "profit":
            profit_total = (profit_total or 0.0) + amount
        else:
            continue
        if not currency and len(row) > 3:
            currency = _norm(row[3]).upper()

    totals.open_value = value_total
    totals.open_profit = profit_total
    return totals, currency


def parse_open_positions(
    rows: list[tuple],
) -> tuple[list[XtbOpenPosition], XtbReportedTotals, str]:
    """Fold the nested open-positions table into one entry per instrument.

    Each instrument occupies an aggregate row followed by one row per lot, and
    both carry the same ticker, volume and value. Reading every row doubles the
    portfolio.

    Aggregates win: for a multi-lot holding the aggregate already carries the
    volume-weighted open price. But the lots still have to be read — they are
    the only rows that report a current price, and they hold the position IDs.
    """
    header_index, columns = _find_header(rows, _OPEN_REQUIRED, SHEET_OPEN)
    totals, currency = _parse_open_summary(rows, header_index)

    positions: dict[str, XtbOpenPosition] = {}
    orphan_lots: dict[str, list[XtbLot]] = {}

    for row in rows[header_index + 1:]:
        symbol = _norm(_cell(row, columns, "ticker")).upper()
        if not symbol:
            continue

        category = _norm(_cell(row, columns, "category"))
        side = _norm(_cell(row, columns, "type"))

        if category and not side:
            # Aggregate row: one per instrument.
            positions[symbol] = XtbOpenPosition(
                xtb_symbol=symbol,
                yahoo_symbol=to_yahoo(symbol),
                instrument=_norm(_cell(row, columns, "instrument/position"))
                or _norm(_cell(row, columns, "instrument")),
                asset_class=category.upper(),
                product=_norm(_cell(row, columns, "product")),
                volume=_to_float(_cell(row, columns, "volume")) or 0.0,
                open_price=_to_float(_cell(row, columns, "open price")) or 0.0,
                value=_to_float(_cell(row, columns, "value")) or 0.0,
                net_profit=_to_float(_cell(row, columns, "net profit")),
                net_profit_pct=_to_float(_cell(row, columns, "net profit %")),
            )
            continue

        if not side:
            continue

        lot = XtbLot(
            position_id=_to_id(_cell(row, columns, "instrument/position")),
            side=side.upper(),
            volume=_to_float(_cell(row, columns, "volume")) or 0.0,
            open_price=_to_float(_cell(row, columns, "open price")) or 0.0,
            current_price=_to_float(_cell(row, columns, "current price")),
            value=_to_float(_cell(row, columns, "value")) or 0.0,
            net_profit=_to_float(_cell(row, columns, "net profit")),
            opened_at=excel_serial_to_iso(_cell(row, columns, "open time (utc)")),
        )
        if symbol in positions:
            positions[symbol].lots.append(lot)
        else:
            # A lot with no aggregate above it. Rather than drop a real
            # holding, collect it and synthesise the instrument below.
            orphan_lots.setdefault(symbol, []).append(lot)

    for symbol, lots in orphan_lots.items():
        volume = sum(lot.volume for lot in lots)
        cost = sum(lot.volume * lot.open_price for lot in lots)
        positions[symbol] = XtbOpenPosition(
            xtb_symbol=symbol,
            yahoo_symbol=to_yahoo(symbol),
            volume=volume,
            open_price=(cost / volume) if volume else 0.0,
            value=sum(lot.value for lot in lots),
            net_profit=sum(lot.net_profit or 0.0 for lot in lots) or None,
            lots=lots,
        )

    # The aggregate row leaves Current price blank; the lots below it do not.
    for position in positions.values():
        if position.current_price is None:
            position.current_price = next(
                (lot.current_price for lot in position.lots if lot.current_price is not None),
                None,
            )

    return list(positions.values()), totals, currency


# ── closed positions ─────────────────────────────────────────────────────────

_CLOSED_REQUIRED = {"instrument", "ticker", "volume", "open price", "close price", "profit/loss"}


def parse_closed_positions(rows: list[tuple]) -> tuple[list[XtbClosedPosition], float | None]:
    """Read finished round trips, returning the sheet's own total separately."""
    header_index, columns = _find_header(rows, _CLOSED_REQUIRED, SHEET_CLOSED)

    closed: list[XtbClosedPosition] = []
    reported: float | None = None

    for row in rows[header_index + 1:]:
        if _is_total_row(row, columns, "ticker"):
            reported = _to_float(_cell(row, columns, "profit/loss"))
            continue

        symbol = _norm(_cell(row, columns, "ticker")).upper()
        if not symbol:
            continue

        closed.append(
            XtbClosedPosition(
                xtb_symbol=symbol,
                instrument=_norm(_cell(row, columns, "instrument")),
                asset_class=_norm(_cell(row, columns, "category")).upper(),
                side=_norm(_cell(row, columns, "type")).upper() or "BUY",
                volume=_to_float(_cell(row, columns, "volume")) or 0.0,
                open_price=_to_float(_cell(row, columns, "open price")) or 0.0,
                close_price=_to_float(_cell(row, columns, "close price")) or 0.0,
                opened_at=excel_serial_to_iso(_cell(row, columns, "open time (utc)")),
                closed_at=excel_serial_to_iso(_cell(row, columns, "close time (utc)")),
                profit_loss=_to_float(_cell(row, columns, "profit/loss")) or 0.0,
                purchase_value=_to_float(_cell(row, columns, "purchase value")),
                sale_value=_to_float(_cell(row, columns, "sale value")),
                commission=_to_float(_cell(row, columns, "commission")) or 0.0,
                position_id=_to_id(_cell(row, columns, "position id")),
            )
        )

    return closed, reported


# ── cash operations ──────────────────────────────────────────────────────────

_CASH_REQUIRED = {"type", "instrument", "ticker", "time", "amount"}


def parse_cash_operations(
    rows: list[tuple],
) -> tuple[list[XtbCashOperation], float | None, list[str]]:
    """Read the cash ledger, returning the sheet's own total separately.

    An unrecognised operation type lands in the `other` bucket and is reported
    as a warning rather than dropped: the balance cross-check would flag the
    resulting drift, but only after the number had already been shown.
    """
    header_index, columns = _find_header(rows, _CASH_REQUIRED, SHEET_CASH)

    operations: list[XtbCashOperation] = []
    reported: float | None = None
    unknown: set[str] = set()

    for row in rows[header_index + 1:]:
        if _is_total_row(row, columns, "time"):
            reported = _to_float(_cell(row, columns, "amount"))
            continue

        label = _norm(_cell(row, columns, "type"))
        amount = _to_float(_cell(row, columns, "amount"))
        if not label or amount is None:
            continue

        bucket = CASH_BUCKETS.get(label.lower(), "other")
        if bucket == "other":
            unknown.add(label)

        operations.append(
            XtbCashOperation(
                type=label,
                bucket=bucket,
                xtb_symbol=_norm(_cell(row, columns, "ticker")).upper(),
                asset_class=_norm(_cell(row, columns, "category")).upper(),
                amount=amount,
                at=excel_serial_to_iso(_cell(row, columns, "time")),
                comment=_norm(_cell(row, columns, "comment")),
                position_id=_to_id(_cell(row, columns, "position id")),
            )
        )

    warnings = [
        f"Tipo de movimiento no reconocido: «{label}». Se contabilizó igual, "
        "pero no se clasificó."
        for label in sorted(unknown)
    ]
    return operations, reported, warnings


# ── entry point ──────────────────────────────────────────────────────────────

def _sheet_rows(workbook: Any, name: str) -> list[tuple]:
    if name not in workbook.sheetnames:
        raise XtbParseError(
            f"Falta la hoja «{name}». Exportá el reporte completo desde XTB."
        )
    sheet = workbook[name]
    # XTB writes `<dimension ref="A1"/>` on every sheet — it declares each one
    # as a single cell. In read_only mode openpyxl believes that and yields one
    # row of one column, so the whole report reads as empty. reset_dimensions
    # makes it scan the rows it actually has.
    reset = getattr(sheet, "reset_dimensions", None)
    if reset is not None:
        reset()
    # Rows come back ragged in read_only mode (a blank row is an empty tuple),
    # and max_row is unreliable, so they are materialised by iteration and
    # every column access is bounds-checked in `_cell`.
    return list(sheet.iter_rows(values_only=True))


def parse_workbook(data: bytes, source_file: str = "") -> XtbSnapshot:
    """Parse an XTB export. Raises XtbParseError for anything unreadable."""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except XtbParseError:
        raise
    except Exception as exc:  # noqa: BLE001 - an upload must never leak a traceback
        raise XtbParseError(
            "No se pudo abrir el archivo. ¿Es el Excel que exporta XTB y está completo?"
        ) from exc

    try:
        open_rows = _sheet_rows(workbook, SHEET_OPEN)
        closed_rows = _sheet_rows(workbook, SHEET_CLOSED)
        cash_rows = _sheet_rows(workbook, SHEET_CASH)

        open_positions, totals, currency = parse_open_positions(open_rows)
        closed_positions, reported_realized = parse_closed_positions(closed_rows)
        cash_operations, reported_cash, warnings = parse_cash_operations(cash_rows)
    finally:
        workbook.close()

    if not open_positions and not closed_positions and not cash_operations:
        raise XtbParseError(
            "El archivo no tiene operaciones. ¿Exportaste el rango de fechas correcto?"
        )

    totals.realized_pnl = reported_realized
    totals.cash_balance = reported_cash

    account = _norm(open_rows[0][1]) if open_rows and len(open_rows[0]) > 1 else ""
    generated_at = (
        excel_serial_to_iso(open_rows[2][1])
        if len(open_rows) > 2 and len(open_rows[2]) > 1
        else None
    )

    currency = currency or "USD"
    if currency != "USD":
        # The display filters format everything with a dollar sign, so a
        # non-USD account would render confidently wrong figures.
        warnings.append(
            f"La cuenta está en {currency}, pero los importes se muestran en dólares."
        )

    unmapped = [p.xtb_symbol for p in open_positions if not p.yahoo_symbol]
    if unmapped:
        warnings.append(
            "Sin equivalente en el proveedor de datos: " + ", ".join(sorted(unmapped))
        )

    return XtbSnapshot(
        imported_at=datetime.now().isoformat(timespec="seconds"),
        source_file=source_file,
        account=account,
        currency=currency,
        generated_at=generated_at,
        open_positions=open_positions,
        closed_positions=closed_positions,
        cash_operations=cash_operations,
        reported=totals,
        warnings=warnings,
    )
